import openpyxl
from openpyxl import Workbook
excel_file = Workbook()
#import openpyxl
#from openpyxl import Workbook
excel_file = Workbook()
wb = openpyxl.load_workbook('student.xlsx')
sheets = ['sheet1', 'sheet2', 'sheet3', 'sheet4', 'sheet5']
excel_sheet = excel_file.create_sheet(title='MasterSheet11', index=0)
n=int(input("number of persons:" ))

for g in range(1, n+1):
    print("enter", g, " person information")
    xin = int(input("enter ps number: "))
    yin = input("enter name: ")
    zin = input("enter mailid: ")
    t = 1
    for sheet in sheets:
        sh = wb[sheet]  # Get a sheet from the workbook.
        max_r = sh.max_row
        max_c = sh.max_column
        if t <= 10:
            for r in range(1, max_r + 1):
                if sh.cell(row=r, column=1).value == xin and sh.cell(row=r, column=2).value == yin and sh.cell(row=r, column=3).value == zin:

                    for c in range(1, max_c + 1):
                        if g==1:
                            str1 = 'A' + str(t)
                            str2 = 'B' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = 'E' + str(t)
                            str2 = 'F' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value

        else:
            for r in range(4, max_r + 1):
                if sh.cell(row=r, column=1).value == xin and sh.cell(row=r, column=2).value == yin and sh.cell(row=r,
                                                                                                               column=3).value == zin:
                    for c in range(4, max_c + 1):
                        if g==1:
                            str1 = 'A' + str(t)
                            str2 = 'B' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = 'E' + str(t)
                            str2 = 'F' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
    excel_file.save(filename="final.xlsx")